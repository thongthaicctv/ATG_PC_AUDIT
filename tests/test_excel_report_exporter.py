import tempfile,unittest
from pathlib import Path
from openpyxl import load_workbook
from core.aggregate_database import AggregateDatabase
from core.backup_manager import backup_database,restore_database
from core.csv_exporter import export_csv
from core.csv_importer import import_previews,preview_files
from core.excel_report_exporter import export_company_report
from test_export import sample_result

class ReportTests(unittest.TestCase):
    def test_report_sheets_backup_restore(self):
        with tempfile.TemporaryDirectory() as tmp:
            root=Path(tmp);db=AggregateDatabase(root/"db.sqlite");p=export_csv(sample_result(),root);import_previews(preview_files([p],db),db)
            report=export_company_report(db,root/"report.xlsx","Tester");wb=load_workbook(report);required={"TONG_QUAN","DANH_SACH_MAY","WINDOWS_11","CAN_NANG_CAP","BAN_QUYEN_WINDOWS","BAN_QUYEN_OFFICE","MAC_IP","RAM","O_DIA","LICH_SU_KIEM_TRA","XUNG_DOT_DU_LIEU","NHAT_KY_DONG_BO"};self.assertEqual(set(wb.sheetnames),required);self.assertEqual(wb["DANH_SACH_MAY"]["A4"].font.name,"Times New Roman")
            backup=backup_database(db.path,root);db.path.write_bytes(b"broken");restore_database(backup,db.path);self.assertGreater(db.path.stat().st_size,100)

    def test_invalid_backup_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            import zipfile
            bad=Path(tmp)/"bad.zip"
            with zipfile.ZipFile(bad,"w") as z:z.writestr("other.txt","x")
            with self.assertRaises(ValueError):restore_database(bad,Path(tmp)/"db")

if __name__=="__main__":unittest.main()
