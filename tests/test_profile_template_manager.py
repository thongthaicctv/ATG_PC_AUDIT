import json,tempfile,unittest
from pathlib import Path

from openpyxl import load_workbook

from core.profile_template_manager import HEADERS,export_template,import_template,load_catalog


class ProfileTemplateManagerTests(unittest.TestCase):
    def test_export_and_import_updates_config(self):
        with tempfile.TemporaryDirectory() as tmp:
            root=Path(tmp);config=root/"app_config.json";config.write_text(json.dumps({"app_name":"ATG"}),encoding="utf-8")
            source=root/"template.xlsx";export_template(source,{"machine_types":["Bộ PC","Laptop"],"departments":["HCNS","IT"],"locations":["Tầng 2","Tầng 3"]})
            wb=load_workbook(source);ws=wb.active;self.assertEqual(tuple(x.value for x in ws[1]),HEADERS);ws.cell(3,2).value="Kỹ thuật";wb.save(source);wb.close()
            catalog=import_template(source,config)
            self.assertEqual(catalog["departments"],["HCNS","Kỹ thuật"]);self.assertEqual(load_catalog(config),catalog)

    def test_import_rejects_missing_required_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            from openpyxl import Workbook
            root=Path(tmp);source=root/"bad.xlsx";wb=Workbook();wb.active.append(["Phòng ban"]);wb.save(source)
            with self.assertRaisesRegex(ValueError,"Loại máy tính"):import_template(source,root/"config.json")


if __name__=="__main__":unittest.main()
