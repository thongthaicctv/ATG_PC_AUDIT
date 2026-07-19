import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from core.aggregate_manager import aggregate_folder
from core.export_manager import export_machine_excel, save_complete_record, save_json, sanitize_filename
from models.audit_result import AuditResult


def sample_result(asset="ATG-001", serial="SERIAL-1", ip="10.10.1.20"):
    return AuditResult(
        metadata={"asset_code": asset, "user": "User", "employee_code": "NV001", "department": "IT", "location": "Tầng 2", "auditor": "Tester", "audit_date": "14/07/2026"},
        computer={"computer_name": "PC-01", "serial_number": serial, "uuid": "UUID-1"},
        cpu={"name": "CPU"}, ram_summary={"total_gb": 16}, ram_modules=[{"slot": "DIMM0", "capacity_gb": 16}],
        disks=[{"disk_index": 0, "model": "SSD", "disk_type": "SSD", "partition_style": "GPT", "capacity_gb": 256, "is_system": True}],
        windows={"edition": "Windows 11 Pro"}, windows11={"overall": "ĐỦ ĐIỀU KIỆN CÀI WINDOWS 11", "conditions": []},
        windows_license={"activation_status": "Đã kích hoạt", "license_channel": "Retail", "partial_key": "XXXXX-XXXXX-XXXXX-XXXXX-ABCDE"},
        office_licenses=[{"product_name": "Microsoft 365 Apps", "activation_status": "Đã kích hoạt", "user_email": "secret@example.com", "tenant_id": "SECRET", "sensitive_account_data": True}],
        network_adapters=[{"interface_index": 1, "name": "Ethernet", "mac_address": "AA-BB-CC-DD-EE-FF", "ipv4": ["10.10.1.10"]}], primary_adapter_index=1,
        ip_plan={"planned_ip": ip, "vlan": "10"}, recommendations=["Không có"],
    )


class ExportTests(unittest.TestCase):
    def test_filename_is_sanitized(self):
        self.assertEqual(sanitize_filename('PC:01/ABC'), "PC_01_ABC")

    def test_json_hides_office_account_by_default(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = save_json(sample_result(), Path(tmp) / "audit_result.json")
            text = path.read_text(encoding="utf-8")
            self.assertNotIn("secret@example.com", text); self.assertNotIn("SECRET", text)

    def test_machine_excel_has_required_sheets_and_format(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_machine_excel(sample_result(), Path(tmp) / "audit.xlsx")
            wb = load_workbook(path)
            required = {"Tong_quan", "Phan_cung", "RAM", "O_dia", "Card_mang_IP", "Windows_11", "Ban_quyen_Windows", "Ban_quyen_Office", "Quy_hoach_IP", "Du_lieu_ky_thuat"}
            self.assertEqual(required, set(wb.sheetnames))
            self.assertEqual(wb["Tong_quan"]["A1"].font.name, "Times New Roman")
            self.assertEqual(wb["Tong_quan"].freeze_panes, "A2")

    def test_complete_record_and_aggregate_duplicates(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp); save_complete_record(sample_result(), root); save_complete_record(sample_result(ip="10.10.1.20"), root)
            output = root / "TongHop.xlsx"; found, used, errors = aggregate_folder(root, output, keep_history=True)
            self.assertEqual((found, used, errors), (2, 2, []))
            wb = load_workbook(output); self.assertIn("Danh_sach_may", wb.sheetnames)
            ws = wb["Danh_sach_may"]; headers = [c.value for c in ws[1]]; ip_col = headers.index("IP dự kiến") + 1
            self.assertEqual(ws.cell(2, ip_col).fill.fgColor.rgb[-6:], "FFC7CE")


if __name__ == "__main__": unittest.main()
