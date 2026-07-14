import csv
import json
import logging
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LOG = logging.getLogger(__name__)
FONT_NAME = "Times New Roman"
STATUS_FILLS = {
    "pass": PatternFill("solid", fgColor="C6EFCE"), "warn": PatternFill("solid", fgColor="FFEB9C"),
    "fail": PatternFill("solid", fgColor="FFC7CE"), "none": PatternFill("solid", fgColor="E7E6E6"),
}


def sanitize_filename(value: Any) -> str:
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(value or ""))
    text = re.sub(r"\s+", "_", text).strip(" ._")
    return text[:80] or "UNKNOWN"


def _plain(result) -> Dict[str, Any]:
    source = result.to_dict() if hasattr(result, "to_dict") else dict(result)
    data = json.loads(json.dumps(source, ensure_ascii=False, default=str))
    for license_item in data.get("office_licenses", []):
        license_item.pop("user_email", None); license_item.pop("tenant_id", None)
        if license_item.get("sensitive_account_data"): license_item["account_export"] = "Đã ẩn theo mặc định"
    return data


def save_json(result, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_plain(result), ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path


def _rows_from_dict(data: Dict[str, Any]) -> List[List[Any]]:
    return [[key, _display(value)] for key, value in data.items()] or [["Trạng thái", "Chưa nhập dữ liệu"]]


def _display(value):
    if value is None: return "Không xác định"
    if isinstance(value, bool): return "Có" if value else "Không"
    if isinstance(value, (list, dict)): return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _sheet(ws, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    ws.append(list(headers))
    for row in rows: ws.append([_display(x) for x in row])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions; ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=FONT_NAME, size=13); cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        max_len = max((len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1)), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 45)
    ws.row_dimensions[1].height = 24


def export_machine_excel(result, path: Path) -> Path:
    data = _plain(result); wb = Workbook(); wb.remove(wb.active)
    c, meta = data.get("computer", {}), data.get("metadata", {})
    primary = next((x for x in data.get("network_adapters", []) if x.get("interface_index") == data.get("primary_adapter_index")), {})
    overview = {
        "Mã tài sản": meta.get("asset_code"), "Tên máy": c.get("computer_name"), "Serial": c.get("serial_number"),
        "UUID": c.get("uuid"), "Người sử dụng": meta.get("user"), "Phòng ban": meta.get("department"),
        "Vị trí": meta.get("location"), "Windows": data.get("windows", {}).get("edition"),
        "Windows Activation": data.get("windows_license", {}).get("activation_status"),
        "Office Activation": ", ".join(dict.fromkeys(x.get("activation_status", "") for x in data.get("office_licenses", []))),
        "Windows 11 Readiness": data.get("windows11", {}).get("display_overall", data.get("windows11", {}).get("overall")),
        "MAC chính": primary.get("mac_address"), "IP hiện tại": ", ".join(primary.get("ipv4", [])),
        "Ngày kiểm tra": meta.get("audit_date"), "Người kiểm tra": meta.get("auditor"), "Ghi chú": meta.get("notes"),
    }
    sheets = []
    sheets.append(("Tong_quan", ["Thuộc tính", "Giá trị"], _rows_from_dict(overview)))
    hardware = {**data.get("computer", {}), **data.get("cpu", {}), **data.get("bios", {}), **data.get("security", {})}
    sheets.append(("Phan_cung", ["Thuộc tính", "Giá trị"], _rows_from_dict(hardware)))
    for name, key, headers, fields in [
        ("RAM", "ram_modules", ["Slot", "Bank", "Dung lượng GB", "Tốc độ", "Configured Clock", "Hãng", "Part Number", "Serial"], ["slot", "bank", "capacity_gb", "speed_mhz", "configured_clock_mhz", "manufacturer", "part_number", "serial_number"]),
        ("O_dia", "disks", ["Index", "Model", "Serial", "Bus", "Media", "Loại", "MBR/GPT", "Dung lượng GB", "Trạng thái"], ["disk_index", "model", "serial_number", "bus_type", "media_type", "disk_type", "partition_style", "capacity_gb", "status"]),
        ("Card_mang_IP", "network_adapters", ["Tên", "Loại", "Kết nối", "MAC", "IPv4", "IPv6", "Mask", "Gateway", "DNS", "DHCP", "DHCP Server", "Index"], ["name", "interface_type", "connection_status", "mac_address", "ipv4", "ipv6", "prefix_or_mask", "default_gateway", "dns_servers", "dhcp_enabled", "dhcp_server", "interface_index"]),
        ("Ban_quyen_Office", "office_licenses", ["Sản phẩm", "Trạng thái", "Cơ chế", "Partial Key", "Raw Status", "Ghi chú"], ["product_name", "activation_status", "mechanism", "partial_key", "raw_status", "note"]),
    ]:
        sheets.append((name, headers, [[row.get(f) for f in fields] for row in data.get(key, [])]))
    win11_fields = ["condition", "actual", "required", "status", "note"]
    sheets.append(("Windows_11", ["Điều kiện", "Thực tế", "Yêu cầu", "Trạng thái", "Ghi chú"], [[x.get(f) for f in win11_fields] for x in data.get("windows11", {}).get("conditions", [])]))
    sheets.append(("Ban_quyen_Windows", ["Thuộc tính", "Giá trị"], _rows_from_dict(data.get("windows_license", {}))))
    sheets.append(("Quy_hoach_IP", ["Thuộc tính", "Giá trị"], _rows_from_dict(data.get("ip_plan", {}))))
    technical = {"audited_at": data.get("audited_at"), "errors": data.get("errors"), "recommendations": data.get("recommendations")}
    sheets.append(("Du_lieu_ky_thuat", ["Thuộc tính", "Giá trị"], _rows_from_dict(technical)))
    for name, headers, rows in sheets:
        _sheet(wb.create_sheet(name), headers, rows)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            text = " ".join(str(c.value or "") for c in row).lower()
            fill = STATUS_FILLS["pass"] if any(x in text for x in ("đạt", "đã kích hoạt")) and "không đạt" not in text else STATUS_FILLS["fail"] if any(x in text for x in ("không đạt", "chưa kích hoạt", "hết hạn")) else STATUS_FILLS["warn"] if any(x in text for x in ("chưa xác định", "cần kiểm tra", "grace")) else None
            if fill:
                for cell in row: cell.fill = fill
    path.parent.mkdir(parents=True, exist_ok=True); wb.save(path); return path


def export_network_csv(result, path: Path) -> Path:
    data = _plain(result); fields = ["name", "interface_type", "connection_status", "mac_address", "ipv4", "ipv6", "prefix_or_mask", "default_gateway", "dns_servers", "dhcp_enabled", "dhcp_server", "interface_index"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields); writer.writeheader()
        for row in data.get("network_adapters", []): writer.writerow({k: _display(row.get(k)) for k in fields})
    return path


def export_license_csv(result, path: Path) -> Path:
    data = _plain(result); fields = ["product", "type", "status", "channel", "partial_key", "expiration", "note"]
    rows = [{"product": data.get("windows_license", {}).get("edition"), "type": "Windows", "status": data.get("windows_license", {}).get("activation_status"), "channel": data.get("windows_license", {}).get("license_channel"), "partial_key": data.get("windows_license", {}).get("partial_key"), "expiration": data.get("windows_license", {}).get("expiration"), "note": data.get("windows_license", {}).get("technical_note")}]
    rows += [{"product": x.get("product_name"), "type": "Office", "status": x.get("activation_status"), "channel": x.get("mechanism"), "partial_key": x.get("partial_key"), "expiration": x.get("expiration"), "note": x.get("note")} for x in data.get("office_licenses", [])]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields); writer.writeheader(); writer.writerows(rows)
    return path


def save_complete_record(result, base_dir: Path) -> Path:
    data = _plain(result); c = data.get("computer", {})
    identity = c.get("serial_number") if c.get("serial_number") not in (None, "", "Không xác định") else c.get("uuid") or c.get("computer_name")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{sanitize_filename(c.get('computer_name'))}_{sanitize_filename(identity)}_{stamp}"
    folder = base_dir / base_name
    counter = 2
    while folder.exists():
        folder = base_dir / f"{base_name}_{counter}"; counter += 1
    (folder / "logs").mkdir(parents=True, exist_ok=True)
    save_json(result, folder / "audit_result.json"); export_machine_excel(result, folder / "audit_result.xlsx")
    export_network_csv(result, folder / "audit_network.csv"); export_license_csv(result, folder / "audit_license.csv")
    (folder / "audit_raw.txt").write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    log_source = Path.home() / "ATG_PC_AUDIT" / "logs" / "atg_pc_audit.log"
    if log_source.exists(): shutil.copy2(log_source, folder / "logs" / "audit.log")
    else: (folder / "logs" / "audit.log").write_text("Không có log lỗi.\n", encoding="utf-8")
    return folder
