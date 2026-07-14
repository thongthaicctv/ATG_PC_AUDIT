import json
import logging
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LOG = logging.getLogger(__name__)
RED = PatternFill("solid", fgColor="FFC7CE")
HEADER = PatternFill("solid", fgColor="1F4E78")


def _valid_record(data: Any) -> bool:
    return isinstance(data, dict) and isinstance(data.get("computer"), dict) and isinstance(data.get("metadata", {}), dict)


def load_records(folder: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    records, errors = [], []
    for path in folder.rglob("audit_result.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig"))
            if not _valid_record(data): raise ValueError("Cấu trúc JSON không hợp lệ")
            data["_source_file"] = str(path); records.append(data)
        except Exception as exc:
            message = f"Bỏ qua {path}: {exc}"; errors.append(message); LOG.warning(message)
    return records, errors


def _identity(record: Dict[str, Any]) -> str:
    c = record.get("computer", {})
    for value in (c.get("serial_number"), c.get("uuid"), c.get("computer_name")):
        if value and value != "Không xác định": return str(value).strip().lower()
    return record.get("_source_file", "unknown")


def select_records(records: List[Dict[str, Any]], keep_history: bool) -> List[Dict[str, Any]]:
    if keep_history: return sorted(records, key=lambda x: x.get("audited_at", ""))
    newest = {}
    for record in records:
        key = _identity(record)
        if key not in newest or record.get("audited_at", "") >= newest[key].get("audited_at", ""):
            newest[key] = record
    return list(newest.values())


def _primary(record):
    return next((x for x in record.get("network_adapters", []) if x.get("interface_index") == record.get("primary_adapter_index")), {})


def _join(values):
    return ", ".join(str(x) for x in values if x not in (None, ""))


def _summary(record: Dict[str, Any]) -> Dict[str, Any]:
    c, m, w, wl, plan = record.get("computer", {}), record.get("metadata", {}), record.get("windows", {}), record.get("windows_license", {}), record.get("ip_plan", {})
    primary = _primary(record); office = record.get("office_licenses", [])
    system_disk = next((x for x in record.get("disks", []) if x.get("is_system")), {})
    return {
        "Mã tài sản": m.get("asset_code"), "Tên máy": c.get("computer_name"), "Serial": c.get("serial_number"), "UUID": c.get("uuid"),
        "Người sử dụng": m.get("user"), "Phòng ban": m.get("department"), "Vị trí": m.get("location"), "Windows": w.get("edition"),
        "Windows Activation": wl.get("activation_status"), "Windows License Channel": wl.get("license_channel"),
        "Office Product": _join(dict.fromkeys(x.get("product_name", "") for x in office)), "Office Activation": _join(dict.fromkeys(x.get("activation_status", "") for x in office)),
        "CPU": record.get("cpu", {}).get("name"), "RAM GB": record.get("ram_summary", {}).get("total_gb"),
        "Ổ hệ thống": f"{system_disk.get('disk_type', '')} {system_disk.get('capacity_gb', '')} GB".strip(),
        "TPM": record.get("security", {}).get("tpm_spec_version"), "Secure Boot": record.get("security", {}).get("secure_boot_enabled"),
        "Windows 11 Readiness": record.get("windows11", {}).get("display_overall", record.get("windows11", {}).get("overall")),
        "MAC chính": primary.get("mac_address"), "IP hiện tại": _join(primary.get("ipv4", [])), "VLAN dự kiến": plan.get("vlan"),
        "IP dự kiến": plan.get("planned_ip"), "Switch": plan.get("switch_name"), "Cổng switch": plan.get("switch_port"),
        "Khuyến nghị": _join(record.get("recommendations", [])), "Ngày kiểm tra": m.get("audit_date") or record.get("audited_at"),
        "Nguồn JSON": record.get("_source_file"),
    }


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows: ws.append(row)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions; ws.sheet_view.showGridLines = False
    for cell in ws[1]: cell.font = Font(name="Times New Roman", size=13, bold=True, color="FFFFFF"); cell.fill = HEADER; cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.font = Font(name="Times New Roman", size=13); cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        length = max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(length + 2, 12), 38)


def _mark_duplicates(ws, headers, duplicate_fields):
    indexes = [headers.index(x) + 1 for x in duplicate_fields if x in headers]
    for col in indexes:
        counts = Counter(str(ws.cell(row, col).value or "").strip().lower() for row in range(2, ws.max_row + 1) if str(ws.cell(row, col).value or "").strip())
        for row in range(2, ws.max_row + 1):
            value = str(ws.cell(row, col).value or "").strip().lower()
            if value and counts[value] > 1: ws.cell(row, col).fill = RED


def export_aggregate(records: List[Dict[str, Any]], path: Path, keep_history=False) -> Path:
    chosen = select_records(records, keep_history); summaries = [_summary(x) for x in chosen]
    wb = Workbook(); wb.remove(wb.active)
    headers = ["STT"] + (list(summaries[0].keys()) if summaries else list(_summary({}).keys()))
    rows = [[i] + [summary.get(h) for h in headers[1:]] for i, summary in enumerate(summaries, 1)]
    main = wb.create_sheet("Danh_sach_may"); _write_sheet(main, headers, rows)
    _mark_duplicates(main, headers, ["Mã tài sản", "Serial", "UUID", "MAC chính", "IP dự kiến"])
    sheet_filters = {
        "Windows_11": ["STT", "Mã tài sản", "Tên máy", "Windows 11 Readiness", "TPM", "Secure Boot", "Khuyến nghị"],
        "Can_nang_cap": ["STT", "Mã tài sản", "Tên máy", "CPU", "RAM GB", "Ổ hệ thống", "Khuyến nghị"],
        "Ban_quyen_Windows": ["STT", "Mã tài sản", "Tên máy", "Windows", "Windows Activation", "Windows License Channel"],
        "Ban_quyen_Office": ["STT", "Mã tài sản", "Tên máy", "Office Product", "Office Activation"],
        "MAC_IP": ["STT", "Mã tài sản", "Tên máy", "MAC chính", "IP hiện tại"],
        "Quy_hoach_IP": ["STT", "Mã tài sản", "Tên máy", "VLAN dự kiến", "IP dự kiến", "Switch", "Cổng switch"],
        "Lich_su_kiem_tra": ["STT", "Mã tài sản", "Tên máy", "Serial", "Ngày kiểm tra", "Nguồn JSON"],
    }
    for name, cols in sheet_filters.items():
        indexes = [headers.index(c) for c in cols]
        selected_rows = [[row[i] for i in indexes] for row in rows]
        if name == "Can_nang_cap": selected_rows = [r for r in selected_rows if str(r[-1] or "").strip()]
        ws = wb.create_sheet(name); _write_sheet(ws, cols, selected_rows)
        if name in ("MAC_IP", "Quy_hoach_IP"): _mark_duplicates(ws, cols, ["MAC chính", "IP dự kiến"])
    ram_rows, disk_rows = [], []
    for record in chosen:
        base = _summary(record)
        for item in record.get("ram_modules", []): ram_rows.append([base["Mã tài sản"], base["Tên máy"], item.get("slot"), item.get("capacity_gb"), item.get("speed_mhz"), item.get("manufacturer"), item.get("part_number")])
        for item in record.get("disks", []): disk_rows.append([base["Mã tài sản"], base["Tên máy"], item.get("disk_index"), item.get("model"), item.get("disk_type"), item.get("partition_style"), item.get("capacity_gb"), item.get("is_system")])
    _write_sheet(wb.create_sheet("RAM"), ["Mã tài sản", "Tên máy", "Slot", "Dung lượng GB", "Tốc độ", "Hãng", "Part Number"], ram_rows)
    _write_sheet(wb.create_sheet("O_dia"), ["Mã tài sản", "Tên máy", "Index", "Model", "Loại", "MBR/GPT", "Dung lượng GB", "Ổ hệ thống"], disk_rows)
    path.parent.mkdir(parents=True, exist_ok=True); wb.save(path); return path


def aggregate_folder(source: Path, output: Path, keep_history=False):
    records, errors = load_records(source)
    if not records: raise ValueError("Không tìm thấy audit_result.json hợp lệ trong thư mục đã chọn.")
    export_aggregate(records, output, keep_history=keep_history)
    return len(records), len(select_records(records, keep_history)), errors
