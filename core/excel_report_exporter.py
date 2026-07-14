from datetime import datetime
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from openpyxl.utils import get_column_letter

from core.duplicate_checker import find_duplicates
from core.license_models import mask_device_id

BLUE=PatternFill("solid",fgColor="1F4E78");WHITE="FFFFFF";THIN=Side(style="thin",color="D9E1F2")
LOG=logging.getLogger(__name__)


def _sheet(wb,name,headers,rows,exporter):
    ws=wb.create_sheet(name);end=max(1,len(headers));ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=end);ws.cell(1,1,f"ATG PC AUDIT - {name}");ws.cell(2,1,f"Ngày xuất: {datetime.now():%d/%m/%Y %H:%M:%S} | Người xuất: {exporter}")
    ws.append([]);ws.append(headers)
    for row in rows:ws.append(row)
    ws.freeze_panes="A5";ws.auto_filter.ref=f"A4:{get_column_letter(end)}{max(4,ws.max_row)}";ws.sheet_view.showGridLines=False
    ws.cell(1,1).font=Font(name="Times New Roman",size=16,bold=True,color="1F4E78");ws.cell(2,1).font=Font(name="Times New Roman",size=13,italic=True)
    for c in ws[4]:c.font=Font(name="Times New Roman",size=13,bold=True,color=WHITE);c.fill=BLUE;c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for row in ws.iter_rows(min_row=5):
        text=" ".join(str(c.value or "") for c in row).lower();fill=PatternFill("solid",fgColor="C6EFCE") if any(x in text for x in ("đạt","đã kích hoạt","hoàn thành")) and "không đạt" not in text else PatternFill("solid",fgColor="FFC7CE") if any(x in text for x in ("không đạt","chưa kích hoạt","xung đột")) else PatternFill("solid",fgColor="FFEB9C") if any(x in text for x in ("cần kiểm tra","cảnh báo","hash không")) else None
        for c in row:
            c.font=Font(name="Times New Roman",size=13);c.alignment=Alignment(vertical="top",wrap_text=True);c.border=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
            if fill:c.fill=fill
    for col in range(1,end+1):
        width=max(len(str(ws.cell(r,col).value or "")) for r in range(1,ws.max_row+1));ws.column_dimensions[get_column_letter(col)].width=min(max(width+2,12),38)
    return ws


def export_company_report(database,path:Path,exporter="Quản trị viên",license_result=None):
    LOG.info("Excel export started")
    records=database.current_records();stats=database.stats();conflicts=find_duplicates(records);wb=Workbook();wb.remove(wb.active)
    overview=[("Tổng số máy",stats["total_machines"]),("Tổng số người sử dụng",stats["users"]),("Tổng số phòng ban",stats["departments"]),("Máy đủ Windows 11",stats["win11_pass"]),("Máy không đủ Windows 11",stats["win11_fail"]),("Máy cần kiểm tra thêm",stats["win11_unknown"]),("Windows chưa kích hoạt",stats["windows_unlicensed"]),("Office chưa kích hoạt",stats["office_unlicensed"]),("Máy RAM dưới 8 GB",stats["low_ram"]),("Máy dùng HDD",stats["hdd"]),("Xung đột",len(conflicts))]
    if license_result:
        overview += [("THÔNG TIN PHẦN MỀM",""),("Tên phần mềm","ATG PC AUDIT"),("Phiên bản","1.0.0"),("Đơn vị được cấp phép",license_result.company),("Tên license",license_result.license_name),("Tính năng","Tổng hợp dữ liệu"),("Ngày hết hạn",license_result.expire_date or "Vĩnh viễn"),("DEVICE_ID",mask_device_id(license_result.device_id)),("Ngày xuất báo cáo",datetime.now().strftime("%d/%m/%Y %H:%M"))]
    _sheet(wb,"TONG_QUAN",["Chỉ tiêu","Số lượng"],overview,exporter)
    cols=[("asset_code","Mã tài sản"),("computer_name","Tên máy"),("assigned_user","Người sử dụng"),("department","Phòng ban"),("location","Vị trí"),("manufacturer","Hãng"),("model","Model"),("serial_number","Serial"),("uuid","UUID"),("cpu_name","CPU"),("ram_total_gb","RAM"),("system_disk_media_type","Ổ hệ thống"),("system_disk_size_gb","Dung lượng"),("system_disk_free_gb","Dung lượng trống"),("os_edition","Windows"),("windows_activation_status","Windows Activation"),("windows_license_type","Windows License Type"),("office_product_summary","Office"),("office_activation_summary","Office Activation"),("tpm_version","TPM"),("secure_boot_status","Secure Boot"),("win11_status","Windows 11"),("primary_mac","MAC"),("current_ipv4","IP hiện tại"),("planned_vlan","VLAN dự kiến"),("planned_ipv4","IP dự kiến"),("switch_name","Switch"),("switch_port","Cổng switch"),("deployment_status","Trạng thái triển khai"),("audit_date_display","Ngày kiểm tra"),("auditor","Người kiểm tra"),("recommendations","Khuyến nghị")]
    _sheet(wb,"DANH_SACH_MAY",[x[1] for x in cols],[[r.get(x[0]) for x in cols] for r in records],exporter)
    wincols=[("asset_code","Mã tài sản"),("computer_name","Tên máy"),("cpu_name","CPU"),("ram_total_gb","RAM"),("system_disk_size_gb","Storage"),("tpm_version","TPM"),("firmware_mode","UEFI"),("secure_boot_status","Secure Boot"),("system_disk_partition_style","GPT"),("win11_status","Kết luận"),("win11_block_reasons","Lý do không đạt"),("recommendations","Khuyến nghị")]
    _sheet(wb,"WINDOWS_11",[x[1] for x in wincols],[[r.get(x[0]) for x in wincols] for r in records],exporter)
    upgrades=[r for r in records if float(r.get("ram_total_gb") or 0)<8 or "HDD" in str(r.get("system_disk_media_type")) or r.get("windows_activation_status")!="Đã kích hoạt" or "Đã kích hoạt" not in str(r.get("office_activation_summary")) or r.get("recommendations")]
    _sheet(wb,"CAN_NANG_CAP",["Mã tài sản","Tên máy","Mức độ ưu tiên","Nội dung cần nâng cấp","Ghi chú quản trị","Dự kiến xử lý"],[[r.get("asset_code"),r.get("computer_name"),"Cao" if "KHÔNG ĐỦ" in str(r.get("win11_status")) else "Trung bình",r.get("recommendations"),"",""] for r in upgrades],exporter)
    _sheet(wb,"BAN_QUYEN_WINDOWS",["Mã tài sản","Tên máy","Người dùng","Windows Edition","Activation Status","License Type","License Channel","Partial Key","Permanent","Expiration","Ngày kiểm tra"],[[r.get(k) for k in ("asset_code","computer_name","assigned_user","os_edition","windows_activation_status","windows_license_type","windows_license_channel","windows_partial_key","windows_permanent_activation","windows_expiration","audit_date_display")] for r in records],exporter)
    office=[]
    for r in records:
        import json
        try:items=json.loads(r.get("office_license_details_json") or "[]")
        except:items=[]
        if not items:items=[{}]
        for x in items:office.append([r.get("asset_code"),r.get("computer_name"),r.get("assigned_user"),x.get("product_name") or r.get("office_product_summary"),x.get("version"),x.get("activation_status") or r.get("office_activation_summary"),x.get("mechanism"),x.get("partial_key"),x.get("expiration")])
    _sheet(wb,"BAN_QUYEN_OFFICE",["Mã tài sản","Tên máy","Người dùng","Sản phẩm","Phiên bản","Activation Status","License Type","Partial Key","Expiration"],office,exporter)
    _sheet(wb,"MAC_IP",["Mã tài sản","Tên máy","Phòng ban","Vị trí","Adapter","Loại card","MAC","IPv4","Gateway","DNS","DHCP","VLAN dự kiến","IP dự kiến","Switch","Cổng switch","Ổ cắm mạng","Trạng thái"],[[r.get(k) for k in ("asset_code","computer_name","department","location","primary_adapter_name","primary_adapter_type","primary_mac","current_ipv4","default_gateway","dns_servers","dhcp_enabled","planned_vlan","planned_ipv4","switch_name","switch_port","network_socket","deployment_status")] for r in records],exporter)
    with database.connect() as con:
        ram=[list(x) for x in con.execute("SELECT m.asset_code,m.computer_name,r.slot,r.bank,r.capacity_gb,r.speed_mhz,r.manufacturer,r.part_number,r.serial_number FROM ram_modules r JOIN audits a ON a.id=r.audit_row_id JOIN machines m ON m.id=a.machine_id")]
        disks=[list(x) for x in con.execute("SELECT m.asset_code,m.computer_name,d.disk_index,d.model,d.serial_number,d.media_type,d.bus_type,d.partition_style,d.size_gb,d.is_system_disk FROM disks d JOIN audits a ON a.id=d.audit_row_id JOIN machines m ON m.id=a.machine_id")]
        history=[list(x) for x in con.execute("SELECT m.asset_code,m.computer_name,a.audit_id,a.audit_time_iso,a.auditor,a.app_version,a.imported_at,a.raw_csv_path,a.hash_verified,a.import_warning FROM audits a JOIN machines m ON m.id=a.machine_id")]
        logs=[list(x) for x in con.execute("SELECT imported_at,import_batch_id,file_name,result,audit_id,export_id,message FROM import_history")]
    _sheet(wb,"RAM",["Mã tài sản","Tên máy","Slot","Bank","Dung lượng GB","Tốc độ","Hãng","Part Number","Serial"],ram,exporter)
    _sheet(wb,"O_DIA",["Mã tài sản","Tên máy","Index","Model","Serial","Media","Bus","MBR/GPT","Dung lượng GB","Ổ hệ thống"],disks,exporter)
    _sheet(wb,"LICH_SU_KIEM_TRA",["Mã tài sản","Tên máy","Audit ID","Ngày kiểm tra","Người kiểm tra","App Version","Ngày import","File nguồn","Hash verified","Ghi chú"],history,exporter)
    _sheet(wb,"XUNG_DOT_DU_LIEU",["Loại xung đột","Giá trị trùng","Máy liên quan","Mã tài sản liên quan","Mức độ","Hướng xử lý","Trạng thái"],[[x[k] for k in ("type","value","machines","assets","severity","resolution","status")] for x in conflicts],exporter)
    _sheet(wb,"NHAT_KY_IMPORT",["Thời gian","Batch ID","File","Kết quả","Audit ID","Export ID","Nội dung"],logs,exporter)
    path.parent.mkdir(parents=True,exist_ok=True);wb.save(path);LOG.info("Excel export completed: %s",path.name);return path
