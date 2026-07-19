import json
from pathlib import Path

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from openpyxl.worksheet.table import Table,TableStyleInfo

from core.resource_utils import resource_path
from core.storage_path_manager import atomic_json_write


DEFAULT_CATALOG={
    "machine_types":["Bộ PC","Laptop","Máy tính bảng"],
    "departments":["HCNS-HR","HCNS-CTV","TTMKT","ECOM-HN","ECOM-HCM","Kho Vận"],
    "locations":["Tầng 3 VP-HN","Tầng 2 VP-HN","Tầng 2 VPHCM","Kho NXY"],
}
HEADERS=("Loại máy tính","Phòng ban","Vị trí làm việc","Ghi chú khác")


def _clean(values):
    result=[];seen=set()
    for value in values:
        text=str(value or "").strip()
        key=text.casefold()
        if text and key not in seen:seen.add(key);result.append(text)
    return result


def config_path():return resource_path("config/app_config.json")


def load_catalog(path=None):
    target=Path(path) if path else config_path()
    try:data=json.loads(target.read_text(encoding="utf-8"))
    except Exception:data={}
    saved=data.get("profile_templates") or {}
    return {key:_clean(saved.get(key) or values) for key,values in DEFAULT_CATALOG.items()}


def save_catalog(catalog,path=None):
    target=Path(path) if path else config_path()
    try:data=json.loads(target.read_text(encoding="utf-8"))
    except Exception:data={}
    normalized={key:_clean(catalog.get(key) or []) for key in DEFAULT_CATALOG}
    for key in DEFAULT_CATALOG:
        if not normalized[key]:raise ValueError(f"Danh mục {key} không được để trống.")
    data["profile_templates"]=normalized;atomic_json_write(target,data);return normalized


def export_template(path,catalog=None):
    catalog=catalog or load_catalog();path=Path(path)
    wb=Workbook();ws=wb.active;ws.title="Biểu mẫu 01";ws.freeze_panes="A2";ws.sheet_view.showGridLines=False;ws.sheet_properties.tabColor="1F4E78"
    ws.append(list(HEADERS));reserved_rows=100
    for i in range(reserved_rows):ws.append([catalog["machine_types"][i] if i<len(catalog["machine_types"]) else "",catalog["departments"][i] if i<len(catalog["departments"]) else "",catalog["locations"][i] if i<len(catalog["locations"]) else "",""])
    header_fill=PatternFill("solid",fgColor="1F4E78");grid_side=Side(style="thin",color="7F8C8D");header_side=Side(style="medium",color="17365D")
    for cell in ws[1]:
        cell.font=Font(name="Calibri",size=11,bold=True,color="FFFFFF");cell.fill=header_fill;cell.alignment=Alignment(vertical="center",horizontal="left");cell.border=Border(left=header_side,right=header_side,top=header_side,bottom=header_side)
    for row_index,row in enumerate(ws.iter_rows(min_row=2,max_row=reserved_rows+1,min_col=1,max_col=4),2):
        fill=PatternFill("solid",fgColor="DCE6F1" if row_index%2==0 else "FFFFFF")
        for cell in row:
            cell.font=Font(name="Calibri",size=10,color="1F1F1F");cell.fill=fill;cell.alignment=Alignment(vertical="center",horizontal="left");cell.border=Border(left=grid_side,right=grid_side,top=grid_side,bottom=grid_side)
        ws.row_dimensions[row_index].height=21
    ws.column_dimensions["A"].width=25;ws.column_dimensions["B"].width=28;ws.column_dimensions["C"].width=30;ws.column_dimensions["D"].width=42;ws.row_dimensions[1].height=28
    table=Table(displayName="DanhMucHoSo",ref=f"A1:D{reserved_rows+1}");table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showFirstColumn=False,showLastColumn=False);ws.add_table(table)
    ws.print_title_rows="1:1";ws.print_area=f"A1:D{reserved_rows+1}";ws.page_setup.orientation="landscape";ws.page_setup.fitToWidth=1;ws.sheet_properties.pageSetUpPr.fitToPage=True
    path.parent.mkdir(parents=True,exist_ok=True);wb.save(path);return path


def import_template(path,config=None):
    wb=load_workbook(Path(path),read_only=True,data_only=True)
    try:
        ws=wb[wb.sheetnames[0]]
        header_map={str(cell.value or "").strip().casefold():idx for idx,cell in enumerate(ws[1],1)}
        aliases={"machine_types":["loại máy tính","loai may tinh"],"departments":["phòng ban","phong ban"],"locations":["vị trí làm việc","vi tri lam viec"]}
        columns={}
        for key,names in aliases.items():
            columns[key]=next((header_map[name] for name in names if name in header_map),None)
            if not columns[key]:raise ValueError(f"Không tìm thấy cột '{HEADERS[list(DEFAULT_CATALOG).index(key)]}'.")
        catalog={key:_clean(ws.cell(row,col).value for row in range(2,ws.max_row+1)) for key,col in columns.items()}
    finally:wb.close()
    return save_catalog(catalog,config)
